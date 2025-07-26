from flask import Blueprint, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import pythoncom  # Required for COM initialization in threads
import win32com.client
import os

export = Blueprint('export', __name__)

def convert_excel_to_pdf(input_path, output_path):
    # Fix: Initialize COM in this thread
    pythoncom.CoInitialize()

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    wb = excel.Workbooks.Open(os.path.abspath(input_path))
    wb.ExportAsFixedFormat(0, os.path.abspath(output_path))
    wb.Close(False)
    excel.Quit()

    # Fix: Uninitialize after COM usage
    pythoncom.CoUninitialize()

@export.route('/excel', methods=['POST'])
def export_excel():
    data = request.get_json()

    if not data:
        return {"error": "No ticket data received"}, 400

    # Load Excel template
    template_path = "Invoice.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active

    # Fill data (adjust cell positions according to your template layout)
    ws["A5"] = data.get("title", "")
    ws["B5"] = ", ".join(data.get("seats", [])) if isinstance(data.get("seats"), list) else data.get("seats", "")
    ws["C5"] = f"₹{data.get('price', '')}"
    ws["D5"] = data.get("time", "")
    ws["E5"] = data.get("bookingId", "")

    # Set column widths
    column_widths = [15, 15, 15, 15, 15]  # Customize per column A–E
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Define a thin border style
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    # Apply border to all cells with content
    for row in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=5):
        for cell in row:
            cell.border = thick_border

    # Save filled Excel as a temp file
    filled_excel_path = "invoice.xlsx"
    filled_pdf_path = "invoice.pdf"
    wb.save(filled_excel_path)

    convert_excel_to_pdf(filled_excel_path, filled_pdf_path)

    return send_file(filled_pdf_path, as_attachment=True)
